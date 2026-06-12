"""
Excel XLSX Exporter — professional formatted spreadsheet with openpyxl
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

FIELDS = [
    "business_name", "owner_name", "email", "phone", "website",
    "address", "city", "state", "country",
    "facebook", "instagram", "linkedin", "twitter", "source"
]
HEADERS = [
    "Business Name", "Owner Name", "Email", "Phone", "Website",
    "Address", "City", "State", "Country",
    "Facebook", "Instagram", "LinkedIn", "Twitter", "Source"
]

# Colors
NAVY   = "0D1E38"
GOLD   = "D4AF37"
WHITE  = "FFFFFF"
LIGHT  = "EFF3FA"
BORDER = "1A2E50"


def export_xlsx(leads: list, output_dir: str, filename: str = None) -> str:
    if not filename:
        filename = f"leads_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = os.path.join(output_dir, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "FreelancerX Leads"

    # ── Title row ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:N1")
    title_cell = ws["A1"]
    title_cell.value = f"FreelancerX Lead Report  ·  {len(leads)} Leads  ·  {datetime.now().strftime('%B %d, %Y')}"
    title_cell.font = Font(name="Calibri", bold=True, size=13, color=GOLD)
    title_cell.fill = PatternFill("solid", fgColor=NAVY)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Header row ─────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="162B4F")
    header_font = Font(name="Calibri", bold=True, size=10, color=GOLD)
    thin = Side(style="thin", color=BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_num, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[2].height = 20

    # ── Data rows ──────────────────────────────────────────────────────────
    for row_num, lead in enumerate(leads, 3):
        row_fill = PatternFill("solid", fgColor=(LIGHT if row_num % 2 == 0 else WHITE))
        for col_num, field in enumerate(FIELDS, 1):
            value = lead.get(field, "") or ""
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = Font(name="Calibri", size=9)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            # Hyperlinks for URL-like fields
            if field == "email" and value and "@" in value:
                cell.hyperlink = f"mailto:{value}"
                cell.font = Font(name="Calibri", size=9, color="1A6FD4", underline="single")
            elif field == "website" and value and value.startswith("http"):
                cell.hyperlink = value
                cell.font = Font(name="Calibri", size=9, color="1A6FD4", underline="single")

    # ── Column widths ──────────────────────────────────────────────────────
    col_widths = {1:28, 2:18, 3:30, 4:16, 5:32, 6:28, 7:14, 8:10, 9:12,
                  10:28, 11:28, 12:28, 13:28, 14:12}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Freeze header ──────────────────────────────────────────────────────
    ws.freeze_panes = "A3"

    # ── Auto-filter ────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}2"

    wb.save(path)
    return path
